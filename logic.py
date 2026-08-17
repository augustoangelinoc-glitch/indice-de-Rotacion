# -*- coding: utf-8 -*-
"""
logic.py
--------
Lógica de negocio del "Índice de Rotación de Inventarios PLUZ".

Este módulo NO tiene nada de interfaz (eso vive en app.py). Aquí solo se
leen los archivos Excel que sube el usuario, se limpian, se cruzan y se
calculan los indicadores. Separar la lógica de la interfaz permite:
  - probar los cálculos con datos de ejemplo sin abrir Streamlit,
  - reutilizar las funciones si más adelante se arma otro reporte,
  - que un error de cálculo no se confunda con un error de la app.

Metodología (replicada del análisis original en Excel):
  Índice de Rotación (veces) = Valor de Salidas (valorizado) / Valor Promedio
  Valor Promedio            = (Valor Inicial + Valor Final) / 2
  Rotación Anualizada        = Índice de Rotación * (365 / días del periodo)
  Días de Inventario         = días del periodo / Índice de Rotación

Todo el cálculo se hace a nivel (Almacén, Código), porque un mismo código
puede repetirse en distintos almacenes con comportamientos de rotación muy
distintos (así estaba también en el archivo original).
"""

import io
import unicodedata
import re
from datetime import datetime

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Utilidades generales
# ---------------------------------------------------------------------------

def _normalize(text) -> str:
    """Pasa un texto a minúsculas, sin tildes y sin símbolos, para poder
    comparar nombres de columnas aunque el usuario los escriba distinto
    (ej. 'Código', 'CODIGO', 'Cod.' deben reconocerse como lo mismo)."""
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def _find_column(columns, aliases):
    """Busca dentro de `columns` (nombres reales del Excel) cuál calza mejor
    con alguno de los `aliases` conocidos para un campo lógico."""
    norm_map = {_normalize(c): c for c in columns}
    for alias in aliases:
        na = _normalize(alias)
        if na in norm_map:
            return norm_map[na]
    for alias in aliases:
        na = _normalize(alias)
        for norm_col, real_col in norm_map.items():
            if na and (na in norm_col or norm_col in na):
                return real_col
    return None


def _clean_codigo_series(serie: pd.Series) -> pd.Series:
    """
    Normaliza la columna Código preservando ceros a la izquierda.

    Problema típico: si el Excel original tiene el código como texto
    ("00854"), Excel/pandas a veces lo interpreta como número y el cero
    inicial se pierde ("854"), lo que rompe el cruce entre archivos
    (Inventario, Salidas, Costos) porque "00854" != "854".

    Solución: se lee SIEMPRE como texto (ver `_read_excel_text_codes`) y
    aquí solo se quitan espacios y el ".0" que a veces queda pegado cuando
    Excel guardó la celda como número con decimales.
    """
    s = serie.astype(str).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)  # "854.0" -> "854"
    s = s.replace({"nan": np.nan, "None": np.nan, "": np.nan})
    return s


def _read_excel_text_codes(file, codigo_aliases) -> pd.DataFrame:
    """
    Lee un Excel en dos pasadas:
      1) una lectura rápida solo para detectar cuál columna es el "Código".
      2) una segunda lectura forzando esa columna a texto (dtype=str), para
         que nunca se pierdan ceros a la izquierda ni se generen decimales
         falsos (854 -> 854.0) por culpa del tipo numérico de Excel.
    """
    # Si viene como bytes (subida repetida desde Streamlit), lo volvemos a envolver
    raw = file.read() if hasattr(file, "read") else file
    if isinstance(raw, (bytes, bytearray)):
        buffer = io.BytesIO(raw)
    else:
        buffer = file

    preview = pd.read_excel(buffer, nrows=0)
    preview.columns = [str(c).strip() for c in preview.columns]
    codigo_col = _find_column(preview.columns, codigo_aliases)

    if isinstance(raw, (bytes, bytearray)):
        buffer = io.BytesIO(raw)
    else:
        buffer.seek(0)

    dtype_map = {codigo_col: str} if codigo_col else None
    df = pd.read_excel(buffer, dtype=dtype_map)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ---------------------------------------------------------------------------
# Alias de columnas aceptados (para detectar automáticamente el archivo)
# ---------------------------------------------------------------------------

INVENTARIO_ALIASES = {
    "periodo": ["periodo"],
    "almacen": ["almacen", "cod almacen", "codigo almacen"],
    "codigo": ["codigo", "material", "cod material", "sku", "item"],
    "descripcion": ["descripcion", "desc", "nombre", "detalle"],
    "unidad": ["u medida", "unidad", "u m", "um", "unid"],
    "cantidad": ["sistema qty", "cantidad", "qty", "existencia", "stock", "saldo"],
    "familia": ["familia", "grupo material", "grupo", "categoria", "clase"],
    "costo_unitario": ["costo kardex", "costo unitario kardex", "costo unitario"],
    "valor_inventario": ["valor inventario", "valor"],
}

SALIDAS_ALIASES = {
    "fecha": ["fecha", "fecha salida", "fecha movimiento"],
    "cod_almacen": ["cod almacen", "codigo almacen"],
    "almacen": ["almacen"],
    "codigo": ["material", "codigo", "cod material", "sku", "item"],
    "descripcion": ["descripcion", "desc", "nombre"],
    "familia": ["grupo material", "familia", "grupo", "categoria"],
    "unidad": ["u m", "unidad", "um", "unid"],
    "cantidad_salida": ["unidades", "cantidad", "cantidad salida", "qty"],
    "costo_unitario": ["costo unitario"],
    "valor_salida": ["valor salida", "valor"],
}

COSTOS_ALIASES = {
    "codigo": ["codigo", "material"],
    "costo_unitario": ["costo unitario kardex", "costo unitario"],
    "fuente": ["fuente", "origen"],
}


# ---------------------------------------------------------------------------
# Carga y limpieza de archivos
# ---------------------------------------------------------------------------

def load_inventario_file(file, tipo: str) -> pd.DataFrame:
    """
    Carga un archivo de Inventario (Inicial o Final).
    tipo: 'Inicial' | 'Final' — solo se usa para dejar registro del origen,
    no cambia la forma de leer el archivo.

    Columnas obligatorias: Almacén, Código, Cantidad (Sistema/Qty).
    Columnas opcionales (si faltan, quedan vacías y se completan más
    adelante con `fill_missing_descriptions`): Descripción, Unidad, Familia,
    Costo Unitario, Valor de Inventario (si no viene Valor, se calcula como
    Cantidad × Costo Unitario cuando ambos estén disponibles).
    """
    aliases = INVENTARIO_ALIASES
    df = _read_excel_text_codes(file, aliases["codigo"])
    mapping = {campo: _find_column(df.columns, alist) for campo, alist in aliases.items()}
    rename = {v: k for k, v in mapping.items() if v is not None}
    df = df.rename(columns=rename)

    faltantes = [f for f in ["almacen", "codigo", "cantidad"] if f not in df.columns]
    if faltantes:
        raise ValueError(
            f"No se pudieron identificar columnas obligatorias en Inventario {tipo}: "
            + ", ".join(faltantes)
            + ". Verifica que el archivo tenga columnas de Almacén, Código y Cantidad/Sistema (Qty)."
        )

    for opcional in ["descripcion", "unidad", "familia", "costo_unitario", "valor_inventario"]:
        if opcional not in df.columns:
            df[opcional] = np.nan

    df["codigo"] = _clean_codigo_series(df["codigo"])
    df["almacen"] = df["almacen"].astype(str).str.strip()
    df["descripcion"] = df["descripcion"].astype(str).str.strip()
    df["descripcion"] = df["descripcion"].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    df["familia"] = df["familia"].astype(str).str.strip().replace({"nan": np.nan, "": np.nan})
    df["unidad"] = df["unidad"].astype(str).str.strip().str.upper().replace({"NAN": np.nan, "": np.nan})
    df["cantidad"] = pd.to_numeric(df["cantidad"], errors="coerce").fillna(0)
    df["costo_unitario"] = pd.to_numeric(df["costo_unitario"], errors="coerce")
    df["valor_inventario"] = pd.to_numeric(df["valor_inventario"], errors="coerce")

    # Si no vino "Valor de Inventario" pero sí Cantidad y Costo Unitario, se calcula.
    falta_valor = df["valor_inventario"].isna() & df["costo_unitario"].notna()
    df.loc[falta_valor, "valor_inventario"] = df.loc[falta_valor, "cantidad"] * df.loc[falta_valor, "costo_unitario"]
    df["valor_inventario"] = df["valor_inventario"].fillna(0)

    df["tipo"] = tipo
    df = df.dropna(subset=["codigo"])
    return df[["tipo", "almacen", "codigo", "descripcion", "familia", "unidad",
               "cantidad", "costo_unitario", "valor_inventario"]]


def load_salidas_file(file) -> pd.DataFrame:
    """
    Carga el archivo de Salidas (movimientos de consumo/despacho).
    Columnas obligatorias: Almacén, Código (Material), Fecha, Unidades.
    Si falta "Valor de Salida" pero hay Costo Unitario, se calcula
    Unidades × Costo Unitario.
    """
    aliases = SALIDAS_ALIASES
    df = _read_excel_text_codes(file, aliases["codigo"])
    mapping = {campo: _find_column(df.columns, alist) for campo, alist in aliases.items()}
    rename = {v: k for k, v in mapping.items() if v is not None}
    df = df.rename(columns=rename)

    faltantes = [f for f in ["almacen", "codigo", "fecha", "cantidad_salida"] if f not in df.columns]
    if faltantes:
        raise ValueError(
            "No se pudieron identificar columnas obligatorias en Salidas: "
            + ", ".join(faltantes)
            + ". Verifica que el archivo tenga columnas de Almacén, Material/Código, Fecha y Unidades."
        )

    for opcional in ["descripcion", "familia", "unidad", "costo_unitario", "valor_salida", "cod_almacen"]:
        if opcional not in df.columns:
            df[opcional] = np.nan

    df["codigo"] = _clean_codigo_series(df["codigo"])

    # Resolución del Almacén: se usa el CÓDIGO de almacén (ej. "A1421") como
    # clave de cruce siempre que exista, porque es el mismo formato que usan
    # los archivos de Inventario Inicial/Final. Algunos archivos de Salidas
    # traen además un nombre descriptivo del almacén (ej. "PRINCIPAL PLUZ -
    # SJL") en una columna aparte ("Almacen"); ese nombre NUNCA se usa para
    # cruzar (antes se usaba por error y los almacenes quedaban duplicados:
    # "A1421" por un lado y "PRINCIPAL PLUZ - SJL" por otro sin cruzar entre
    # sí), pero se conserva en "almacen_nombre" solo para mostrarlo bonito
    # en los reportes.
    df["almacen_nombre"] = df["almacen"].astype(str).str.strip().replace({"nan": np.nan, "": np.nan})
    df["cod_almacen"] = df["cod_almacen"].astype(str).str.strip().replace({"nan": np.nan, "": np.nan})
    df["almacen"] = df["cod_almacen"].fillna(df["almacen_nombre"])

    df["descripcion"] = df["descripcion"].astype(str).str.strip().replace({"nan": np.nan, "": np.nan})
    df["familia"] = df["familia"].astype(str).str.strip().replace({"nan": np.nan, "": np.nan})
    df["unidad"] = df["unidad"].astype(str).str.strip().str.upper().replace({"NAN": np.nan, "": np.nan})
    df["fecha_original"] = df["fecha"]
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce", dayfirst=True)
    df["cantidad_salida"] = pd.to_numeric(df["cantidad_salida"], errors="coerce")
    df["costo_unitario"] = pd.to_numeric(df["costo_unitario"], errors="coerce")
    df["valor_salida"] = pd.to_numeric(df["valor_salida"], errors="coerce")

    falta_valor = df["valor_salida"].isna() & df["costo_unitario"].notna()
    df.loc[falta_valor, "valor_salida"] = df.loc[falta_valor, "cantidad_salida"] * df.loc[falta_valor, "costo_unitario"]

    df = df.dropna(subset=["codigo"])
    return df[["almacen", "almacen_nombre", "codigo", "descripcion", "familia", "unidad", "fecha",
               "fecha_original", "cantidad_salida", "costo_unitario", "valor_salida"]]


def load_costos_file(file) -> pd.DataFrame:
    """Carga el archivo opcional de Costos Unitarios (Código -> Costo).
    Si el usuario no lo sube, el costo se obtiene automáticamente de los
    inventarios (ver `build_costos_unitarios`)."""
    aliases = COSTOS_ALIASES
    df = _read_excel_text_codes(file, aliases["codigo"])
    mapping = {campo: _find_column(df.columns, alist) for campo, alist in aliases.items()}
    rename = {v: k for k, v in mapping.items() if v is not None}
    df = df.rename(columns=rename)
    if "codigo" not in df.columns or "costo_unitario" not in df.columns:
        raise ValueError(
            "El archivo de Costos Unitarios debe tener columnas de Código y Costo Unitario."
        )
    df["codigo"] = _clean_codigo_series(df["codigo"])
    df["costo_unitario"] = pd.to_numeric(df["costo_unitario"], errors="coerce")
    df = df.dropna(subset=["codigo"])
    return df[["codigo", "costo_unitario"]].drop_duplicates(subset=["codigo"])


# ---------------------------------------------------------------------------
# Descripciones, familias y unidades siempre completas
# ---------------------------------------------------------------------------

def build_master_lookup(*dataframes) -> pd.DataFrame:
    """
    Construye un maestro código -> (descripción, familia, unidad) combinando
    TODAS las fuentes cargadas (Inventario Inicial, Final y Salidas).

    Esto es clave para que "siempre vaya la descripción del código": si un
    material no tiene descripción en el Inventario Final pero sí en el
    Inicial o en Salidas, igual aparecerá con su nombre en los reportes,
    en vez de mostrar un código pelado.
    """
    piezas = []
    for df in dataframes:
        if df is None or df.empty:
            continue
        cols = [c for c in ["codigo", "descripcion", "familia", "unidad"] if c in df.columns]
        piezas.append(df[cols])
    todo = pd.concat(piezas, ignore_index=True)

    def _first_valid(serie):
        serie = serie.dropna()
        return serie.iloc[0] if not serie.empty else np.nan

    maestro = todo.groupby("codigo", as_index=False).agg(
        descripcion=("descripcion", _first_valid),
        familia=("familia", _first_valid) if "familia" in todo.columns else ("codigo", "first"),
        unidad=("unidad", _first_valid) if "unidad" in todo.columns else ("codigo", "first"),
    )
    return maestro


def fill_missing_descriptions(df: pd.DataFrame, maestro: pd.DataFrame) -> pd.DataFrame:
    """Rellena descripción/familia/unidad faltantes en `df` usando el maestro.
    Si de todas formas no se encuentra, deja un texto explícito en vez de
    un vacío, para que sea evidente en el reporte que falta información."""
    df = df.merge(maestro, on="codigo", how="left", suffixes=("", "_maestro"))
    for campo in ["descripcion", "familia", "unidad"]:
        col_maestro = f"{campo}_maestro"
        if col_maestro in df.columns:
            df[campo] = df[campo].fillna(df[col_maestro])
            df = df.drop(columns=[col_maestro])
    df["descripcion"] = df["descripcion"].fillna("(Sin descripción registrada)")
    df["familia"] = df["familia"].fillna("(Sin familia)")
    df["unidad"] = df["unidad"].fillna("(Sin unidad)")
    return df


# ---------------------------------------------------------------------------
# Costo unitario por código (si no se sube el archivo de Costos_Unitarios)
# ---------------------------------------------------------------------------

def build_costos_unitarios(inv_final: pd.DataFrame, inv_inicial: pd.DataFrame,
                            costos_manual: pd.DataFrame = None) -> pd.DataFrame:
    """
    Determina el costo unitario Kardex de cada código con esta prioridad:
      1) Archivo de Costos_Unitarios subido manualmente (si el usuario lo sube).
      2) Costo Kardex del Inventario Final.
      3) Costo Kardex del Inventario Inicial (si el material no está en el Final).
    Devuelve columnas: codigo, costo_unitario, fuente.
    """
    final_c = inv_final[["codigo", "costo_unitario"]].dropna(subset=["costo_unitario"])
    final_c = final_c[final_c["costo_unitario"] > 0].drop_duplicates(subset=["codigo"])
    final_c["fuente"] = "Inv. Final"

    inicial_c = inv_inicial[["codigo", "costo_unitario"]].dropna(subset=["costo_unitario"])
    inicial_c = inicial_c[inicial_c["costo_unitario"] > 0].drop_duplicates(subset=["codigo"])
    inicial_c["fuente"] = "Inv. Inicial"

    combinado = pd.concat([final_c, inicial_c], ignore_index=True).drop_duplicates(subset=["codigo"], keep="first")

    if costos_manual is not None and not costos_manual.empty:
        manual = costos_manual.copy()
        manual["fuente"] = "Archivo Costos_Unitarios"
        combinado = pd.concat([manual, combinado], ignore_index=True).drop_duplicates(subset=["codigo"], keep="first")

    return combinado[["codigo", "costo_unitario", "fuente"]]


# ---------------------------------------------------------------------------
# Cálculo del índice de rotación
# ---------------------------------------------------------------------------

def compute_rotacion(
    inv_inicial: pd.DataFrame,
    inv_final: pd.DataFrame,
    salidas: pd.DataFrame,
    costos: pd.DataFrame,
    maestro: pd.DataFrame,
    dias_periodo: int,
    umbral_rotacion_alta: float = 4.0,
) -> pd.DataFrame:
    """
    Calcula el índice de rotación por (Almacén, Código).

    umbral_rotacion_alta: rotación ANUALIZADA mínima (en veces) para
    clasificar un material como "Rotación alta". Es un parámetro editable
    en la app porque el umbral "correcto" depende del rubro/sector.
    """
    # Valorizar las salidas que no traían costo unitario propio.
    salidas = salidas.copy()
    sin_costo_propio = salidas["valor_salida"].isna() | (salidas["cantidad_salida"].notna() & salidas["valor_salida"] == 0)
    salidas = salidas.merge(costos, on="codigo", how="left", suffixes=("", "_ref"))
    usar_ref = salidas["valor_salida"].isna() & salidas["costo_unitario_ref"].notna()
    salidas.loc[usar_ref, "valor_salida"] = salidas.loc[usar_ref, "cantidad_salida"] * salidas.loc[usar_ref, "costo_unitario_ref"]
    salidas["sin_costo_disponible"] = salidas["valor_salida"].isna()
    salidas["valor_salida"] = salidas["valor_salida"].fillna(0)

    agg_salidas = salidas.groupby(["almacen", "codigo"], as_index=False).agg(
        cantidad_salidas=("cantidad_salida", "sum"),
        valor_salidas=("valor_salida", "sum"),
        num_movimientos=("cantidad_salida", "count"),
        movimientos_sin_costo=("sin_costo_disponible", "sum"),
    )

    # Mapa código-de-almacén -> nombre descriptivo (si el archivo de Salidas lo trae),
    # solo para mostrar en pantalla; el cruce siempre se hace por código.
    if "almacen_nombre" in salidas.columns:
        nombres_almacen = (
            salidas.dropna(subset=["almacen_nombre"])
            .drop_duplicates(subset=["almacen"])[["almacen", "almacen_nombre"]]
        )
    else:
        nombres_almacen = pd.DataFrame(columns=["almacen", "almacen_nombre"])

    ini = inv_inicial.groupby(["almacen", "codigo"], as_index=False).agg(
        cantidad_inicial=("cantidad", "sum"), valor_inicial=("valor_inventario", "sum")
    )
    fin = inv_final.groupby(["almacen", "codigo"], as_index=False).agg(
        cantidad_final=("cantidad", "sum"), valor_final=("valor_inventario", "sum")
    )

    base = ini.merge(fin, on=["almacen", "codigo"], how="outer")
    base = base.merge(agg_salidas, on=["almacen", "codigo"], how="outer")
    for c in ["cantidad_inicial", "valor_inicial", "cantidad_final", "valor_final",
              "cantidad_salidas", "valor_salidas", "num_movimientos", "movimientos_sin_costo"]:
        base[c] = base[c].fillna(0)

    base["valor_promedio"] = (base["valor_inicial"] + base["valor_final"]) / 2

    base["indice_rotacion"] = np.where(
        base["valor_promedio"] > 0, base["valor_salidas"] / base["valor_promedio"], np.nan
    )
    base["rotacion_anualizada"] = base["indice_rotacion"] * (365.0 / max(dias_periodo, 1))
    base["dias_inventario"] = np.where(
        base["indice_rotacion"] > 0, dias_periodo / base["indice_rotacion"], np.nan
    )

    def clasificar(row):
        if row["cantidad_salidas"] <= 0:
            return "Sin rotación"
        if pd.isna(row["rotacion_anualizada"]):
            return "Sin datos de inventario"
        return "Rotación alta" if row["rotacion_anualizada"] >= umbral_rotacion_alta else "Rotación baja"

    base["clasificacion"] = base.apply(clasificar, axis=1)

    base = fill_missing_descriptions(base, maestro)
    base = base.merge(nombres_almacen, on="almacen", how="left")
    base["almacen_nombre"] = base["almacen_nombre"].fillna(base["almacen"])
    base["key"] = base["almacen"] + "|" + base["codigo"]

    cols = [
        "almacen", "almacen_nombre", "codigo", "descripcion", "familia", "unidad",
        "cantidad_inicial", "cantidad_final", "cantidad_salidas",
        "valor_inicial", "valor_final", "valor_promedio", "valor_salidas",
        "indice_rotacion", "rotacion_anualizada", "dias_inventario",
        "clasificacion", "num_movimientos", "movimientos_sin_costo", "key",
    ]
    return base[cols].sort_values(["almacen", "valor_salidas"], ascending=[True, False]).reset_index(drop=True)


def compute_materiales_sin_rotacion(rotacion_df: pd.DataFrame) -> pd.DataFrame:
    """Materiales con inventario (inicial o final > 0) pero SIN ninguna
    salida en el periodo: dinero inmovilizado en stock que no se mueve."""
    df = rotacion_df[
        (rotacion_df["cantidad_salidas"] <= 0)
        & ((rotacion_df["cantidad_inicial"] > 0) | (rotacion_df["cantidad_final"] > 0))
    ].copy()
    total = df["valor_promedio"].sum()
    df["pct_del_total_inmovilizado"] = np.where(total > 0, df["valor_promedio"] / total, 0)
    cols = [
        "almacen", "almacen_nombre", "codigo", "descripcion", "familia", "unidad",
        "cantidad_inicial", "cantidad_final", "valor_promedio", "pct_del_total_inmovilizado",
    ]
    return df[cols].sort_values("valor_promedio", ascending=False).reset_index(drop=True)


def compute_abc(rotacion_df: pd.DataFrame, umbral_a: float = 0.80, umbral_b: float = 0.95) -> pd.DataFrame:
    """
    Clasificación ABC por valor de salidas (Pareto), por (Almacén, Código).
    umbral_a / umbral_b: cortes acumulados por defecto 80% / 95% (ajustables).
    """
    df = rotacion_df[rotacion_df["valor_salidas"] > 0].copy()
    df = df.sort_values("valor_salidas", ascending=False).reset_index(drop=True)
    total = df["valor_salidas"].sum()
    df["pct_individual"] = df["valor_salidas"] / total if total > 0 else 0
    df["pct_acumulado"] = df["pct_individual"].cumsum()
    df["ranking"] = np.arange(1, len(df) + 1)

    def clasificar(pct_acum):
        if pct_acum <= umbral_a:
            return "A"
        if pct_acum <= umbral_b:
            return "B"
        return "C"

    df["clase_abc"] = df["pct_acumulado"].apply(clasificar)
    cols = ["ranking", "almacen", "almacen_nombre", "codigo", "descripcion", "valor_salidas",
            "pct_individual", "pct_acumulado", "clase_abc"]
    return df[cols]


def build_resumen(rotacion_df: pd.DataFrame, salidas_df: pd.DataFrame,
                   fecha_inicio, fecha_fin, dias_periodo: int) -> dict:
    """Arma el diccionario de indicadores generales (equivalente a la hoja
    'Resumen' del Excel original) para mostrar como KPIs en el dashboard."""
    por_almacen = rotacion_df.groupby(["almacen", "almacen_nombre"], as_index=False).agg(
        inv_inicial=("valor_inicial", "sum"),
        inv_final=("valor_final", "sum"),
        inv_promedio=("valor_promedio", "sum"),
        valor_salidas=("valor_salidas", "sum"),
        num_movimientos=("num_movimientos", "sum"),
        movimientos_sin_costo=("movimientos_sin_costo", "sum"),
    )
    por_almacen["indice_rotacion"] = np.where(
        por_almacen["inv_promedio"] > 0, por_almacen["valor_salidas"] / por_almacen["inv_promedio"], np.nan
    )
    por_almacen["rotacion_anualizada"] = por_almacen["indice_rotacion"] * (365.0 / max(dias_periodo, 1))
    por_almacen["dias_inventario"] = np.where(
        por_almacen["indice_rotacion"] > 0, dias_periodo / por_almacen["indice_rotacion"], np.nan
    )

    total_inv_prom = por_almacen["inv_promedio"].sum()
    total_salidas = por_almacen["valor_salidas"].sum()
    indice_total = total_salidas / total_inv_prom if total_inv_prom > 0 else np.nan

    return {
        "por_almacen": por_almacen,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "dias_periodo": dias_periodo,
        "indice_total": indice_total,
        "rotacion_anualizada_total": indice_total * (365.0 / max(dias_periodo, 1)) if pd.notna(indice_total) else np.nan,
        "dias_inventario_total": dias_periodo / indice_total if indice_total and indice_total > 0 else np.nan,
        "total_inv_inicial": por_almacen["inv_inicial"].sum(),
        "total_inv_final": por_almacen["inv_final"].sum(),
        "total_inv_promedio": total_inv_prom,
        "total_valor_salidas": total_salidas,
    }


# ---------------------------------------------------------------------------
# Exportación a Excel (mismo formato que el archivo original)
# ---------------------------------------------------------------------------

def build_export_excel(rotacion_df, sin_rotacion_df, abc_df, resumen, output_path):
    """Genera el Excel de resultados con las mismas hojas del análisis
    original (Resumen, Rotación_por_Material, Materiales_Sin_Rotación,
    ABC_Salidas), con encabezados con color y columnas autoajustadas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, 10), 45)
        ws.freeze_panes = "A2"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        resumen_rows = [
            ["Periodo analizado", f"{resumen['fecha_inicio']} a {resumen['fecha_fin']} ({resumen['dias_periodo']} días)"],
            ["Inventario inicial total (S/.)", round(resumen["total_inv_inicial"], 2)],
            ["Inventario final total (S/.)", round(resumen["total_inv_final"], 2)],
            ["Inventario promedio total (S/.)", round(resumen["total_inv_promedio"], 2)],
            ["Valor de salidas total (S/.)", round(resumen["total_valor_salidas"], 2)],
            ["Índice de rotación total (periodo, veces)", round(resumen["indice_total"], 4) if pd.notna(resumen["indice_total"]) else "N/D"],
            ["Rotación anualizada total (veces)", round(resumen["rotacion_anualizada_total"], 4) if pd.notna(resumen["rotacion_anualizada_total"]) else "N/D"],
            ["Días de inventario total", round(resumen["dias_inventario_total"], 1) if pd.notna(resumen["dias_inventario_total"]) else "N/D"],
        ]
        pd.DataFrame(resumen_rows, columns=["Indicador", "Valor"]).to_excel(writer, sheet_name="Resumen", index=False)
        resumen["por_almacen"].round(2).to_excel(writer, sheet_name="Resumen", index=False, startrow=len(resumen_rows) + 3)

        rotacion_df.rename(columns={
            "almacen": "Cód. Almacén", "almacen_nombre": "Almacén", "codigo": "Código", "descripcion": "Descripción", "familia": "Familia",
            "unidad": "U.M.", "cantidad_inicial": "Cantidad Inicial", "cantidad_final": "Cantidad Final",
            "cantidad_salidas": "Cantidad de Salidas", "valor_inicial": "Valor Inicial (S/.)",
            "valor_final": "Valor Final (S/.)", "valor_promedio": "Valor Promedio (S/.)",
            "valor_salidas": "Valor de Salidas (S/.)", "indice_rotacion": "Índice de Rotación (veces)",
            "rotacion_anualizada": "Rotación Anualizada (veces)", "dias_inventario": "Días Inventario",
            "clasificacion": "Clasificación", "num_movimientos": "N° Movimientos",
            "movimientos_sin_costo": "Movimientos sin Costo", "key": "Key",
        }).round(2).to_excel(writer, sheet_name="Rotacion_por_Material", index=False)

        sin_rotacion_df.rename(columns={
            "almacen": "Cód. Almacén", "almacen_nombre": "Almacén", "codigo": "Código", "descripcion": "Descripción", "familia": "Familia",
            "unidad": "U.M.", "cantidad_inicial": "Cantidad Inicial", "cantidad_final": "Cantidad Final",
            "valor_promedio": "Valor Promedio Inmovilizado (S/.)", "pct_del_total_inmovilizado": "% del Total Inmovilizado",
        }).round(4).to_excel(writer, sheet_name="Materiales_Sin_Rotacion", index=False)

        abc_df.rename(columns={
            "ranking": "Ranking", "almacen": "Cód. Almacén", "almacen_nombre": "Almacén", "codigo": "Código", "descripcion": "Descripción",
            "valor_salidas": "Valor Salidas (S/.)", "pct_individual": "% Individual",
            "pct_acumulado": "% Acumulado", "clase_abc": "Clase ABC",
        }).round(4).to_excel(writer, sheet_name="ABC_Salidas", index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            if ws.max_row >= 1 and ws.max_column >= 1:
                try:
                    style_header(ws)
                except Exception:
                    pass

    return output_path
